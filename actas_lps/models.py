import pandas as pd
from openpyxl import load_workbook
from django.db import models
from datetime import datetime
import os
import pytz
import zipfile

pd.options.display.float_format = "{:.2f}".format


class actas_bd(models.Model):
    """
    Model representing actas_bd table in the database.
    """
    eecc = models.CharField(max_length=100)
    project = models.CharField(max_length=100)
    oc = models.BigIntegerField()
    ip_hijo = models.CharField(max_length=100)
    total_OC = models.FloatField()
    total_certificar = models.FloatField()
    termino_obra = models.DateField()
    servicio_obra = models.DateField()
    posiciones = models.CharField(max_length=100)
    date_created = models.DateTimeField(auto_now_add=True)

    def __str__(self):
        """
        Returns a string representation of the object.
        """
        return (f"{self.eecc} - {self.project} - {self.oc} - {self.ip_hijo} - {self.total_OC} - {self.total_certificar} - {self.termino_obra} - {self.servicio_obra} - {self.posiciones} - {self.date_created}")


class csv_files(models.Model):
    """
    Model representing csv_files table in the database.
    """
    file = models.FileField(upload_to='actas_lps/csv_files/')

    def __str__(self):
        """
        Returns a string representation of the object.
        """
        return self.file.name
    
    def delete(self, *args, **kwargs):
        """
        Deletes the file and the object from the database.
        """
        self.file.delete()
        super().delete(*args, **kwargs)
    
    def m_read_csv(self):
        """
        Reads the CSV file and returns a pandas DataFrame.
        """
        df = pd.read_csv(self.file.path , sep=';', encoding='latin-1')
        return df
    

class clean_actas(models.Model):
    """
    Model representing clean_actas table in the database.
    """
    def __init__(self, file_csv):
        """
        Initializes the clean_actas object with a CSV file.
        """
        self.file_csv = file_csv

    def no_spaces(self, column):
        """
        Removes white spaces from the specified column in the CSV file.
        """
        if self.file_csv[column].dtype == 'object':
            self.file_csv[column] = self.file_csv[column].str.strip()
            return self.file_csv[column]
        else:
            return self.file_csv[column]

    def no_commas(self, column):
        """
        Removes commas from the specified column in the CSV file.
        """
        if self.file_csv[column].dtype == 'object':
            self.file_csv[column] = self.file_csv[column].str.replace(',', '').astype(float)
            return self.file_csv[column]
        else:
            return self.file_csv[column]
    
    def convert_to_int(self, column):
        """
        Converts the specified column to integer in the CSV file.
        """
        if self.file_csv[column].dtype == 'object' or self.file_csv[column].dtype == 'float':
            self.file_csv[column] = self.file_csv[column].astype(int)
            return self.file_csv[column]
        else:
            return self.file_csv[column]
    
    def value_rounded_2(self, columns):
        """
        Rounds the values in the specified columns to 2 decimals in the CSV file.
        """
        for column in columns:
            self.file_csv[column] = self.file_csv[column].round(2)
        return self.file_csv[column]

    def clean_nan(self, column):
        """
        Cleans the NaN values in the specified column of the CSV file.
        """
        if self.file_csv[column].isnull().values.any():
            self.file_csv[column].fillna('', inplace=True)
            return self.file_csv[column]
        else:
            return self.file_csv[column]
    
    def upper_case_column(self, column):
        """
        Converts the values in the specified column to uppercase in the CSV file.
        """
        self.file_csv[column] = self.file_csv[column].str.upper()
        return self.file_csv[column]
    
    def return_df(self):
        """
        Returns the cleaned CSV file as a pandas DataFrame.
        """
        return self.file_csv
    

class Print(models.Model):
    """
    Model representing Print table in the database.
    """
    sheets={'Proyecto':'C7', 'OC':'H10', 'EECC':'C8', 'total_OC':'C9', 'total_certificar':'H9', 'termino_obra':'E18' ,'servicio_obra':'E19','posiciones':'H8'}
    sheetslist = [key for key in sheets]
    date_actas_created = datetime.now(pytz.timezone('America/Lima')).strftime('%d.%m')

    def __init__(self, df):
        """
        Initializes the Print object with a pandas DataFrame.
        """
        self.df = df

    def print_actas(self):
        """
        Prints actas based on the values in the pandas DataFrame.
        """
        text1 = 'ACTA ACEPTACION FINAL'
        text2 = 'ACTA ACEPTACION PARCIAL'
        count = 0
        wb = load_workbook(r'actas_lps/printed_xls/Plantilla_ActaPangeaco.xlsx')
        ws = wb.active
        for i in range (len(self.df)):
            for sn in self.sheetslist:
                ws[self.sheets[sn]] = self.df.loc[i,sn]
                if "OC" == sn:
                    oc_value = self.df.loc[i,sn]
                elif "EECC" == sn:
                    eecc_value = self.df.loc[i,sn]
                elif "Proyecto" == sn:
                    proy_value = self.df.loc[i,sn]
                elif "total_OC" == sn:
                    t_oc = self.df.loc[i,sn]
                elif "total_certificar" == sn:
                    t_oc_c = self.df.loc[i,sn]
            if t_oc == t_oc_c:
                ws['D4'] = "ACTA DE ACEPTACIÓN FINAL - PANGEA"
                wb.save(f'actas_lps/printed_xls/printed_actas/{text1} - {eecc_value} - {oc_value} - {proy_value} - {self.date_actas_created}.xlsx')
            elif t_oc != t_oc_c:
                ws['D4'] = "ACTA DE ACEPTACIÓN PARCIAL - PANGEA"
                wb.save(f'actas_lps/printed_xls/printed_actas/{text2} - {eecc_value} - {oc_value} - {proy_value} - {self.date_actas_created}.xlsx')
            count += 1
        return 'Se crearon {} archivos en total'.format(count)

    def delete_files(self, directory):
        """
        Deletes all files in the specified directory.
        """
        for filename in os.listdir(directory):
            file_path = os.path.join(directory, filename)
            try:
                if os.path.isfile(file_path):
                    os.unlink(file_path)
            except Exception as e:
                print('Failed to delete %s. Reason: %s' % (file_path, e))

    def created_excel_files_zip(self):
        """
        Creates a zip file containing all the Excel files in a directory.
        """
        directory_xlsx = r'actas_lps/printed_xls/printed_actas/'
        files = [f for f in os.listdir(directory_xlsx) if f.endswith('.xlsx')]
        directory_zip = r'actas_lps/printed_xls/zip_files/'
        with zipfile.ZipFile(os.path.join(directory_zip, 'excel_files.zip'), 'w') as zipf:
            for file in files:
                zipf.write(os.path.join(directory_xlsx, file), arcname=file)
        with open(os.path.join(directory_zip, 'excel_files.zip'), 'rb') as f:
            data = f.read()
        self.delete_files(directory_xlsx)
        return data
    